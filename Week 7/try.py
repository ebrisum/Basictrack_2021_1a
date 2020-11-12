import pandas as pd
import datetime
import openpyxl
from openpyxl.utils.cell import get_column_letter
from collections import defaultdict

def invoice_generator():
    wb = openpyxl.load_workbook(r'C:\Users\warne\Downloads\Invoice-Template-top.xlsx')
    sheet = wb.active
    number_rows = sheet.max_row
    number_columns = sheet.max_column

    replacement1 = {
        "company": company,
        "date": date,
        "address": address,
        "id": "#" + str(new_id)
    }

    for i in range(number_columns):
        for k in range(number_rows):
            cell = str(sheet[get_column_letter(i+1)+str(k+1)].value)
            for key in replacement1.keys():
                if str(cell) == key:
                    new_cell2 = replacement1.get(key)
                    sheet[get_column_letter(i+1)+str(k+1)] = str(new_cell2)

    current_row = 16
    current_column = 2
    for eachRow in replacement_2.keys():
        for i in range(len(all_items)):
            new_cell = replacement_2.get(eachRow)[i]
            print(new_cell)
            sheet.cell(row=current_row, column=current_column).value = new_cell
            current_row += 1
        current_column += 1
        current_row -= int(len(all_items))
    wb.save(str(company) + "#" + str(new_id) + ".xlsx")

# Load in data as pandas DataFrame:
company_address = pd.read_csv("company_address.csv", delimiter=",")
df_address = pd.DataFrame(company_address)
invoices = pd.read_csv("invoices.csv", delimiter=",")
df_invoices = pd.DataFrame(invoices)

# generate ID:
if len(df_invoices) > 1:
    new_id = df_invoices.iloc[-1, 0] + 1
    print(new_id)
else:
    new_id = 1

# Check if company is in database
company = input("Company: ")
date = datetime.datetime.today().strftime('%d-%m-%Y')
time = datetime.datetime.today().strftime('%H:%M:%S')

if company not in df_address['Company'].values:
    address = input("Addres: ")
    new_company = {
        "Company": company,
        "Address": address
    }
    with open('company_address.csv', 'a') as company_file:
        company_file.write(new_company.get("Company") + "," + new_company.get("Address") + "\n")
else:
    find_address = df_address.loc[df_address['Company'] == company]
    address = find_address.iloc[0, 1]

# Inputs:
all_items = input("Amount of items: ")

total = 0
replacement_2 = defaultdict(list)
inputs = ["Product description: ", "Quantity: ", "Unit price: ", "VAT-rate: " ]
variables = ["DESCRIPTION", "QTY", "UNIT PRICE", "VAT-rate"]

for i in range(int(all_items)):
    for x, y in zip(variables, inputs):
        replacement_2[x] = list()
        replacement_2[x].append(input(y))
    subtotal = (float(replacement_2.get("QTY")[i]) * float(replacement_2.get("UNIT PRICE")[i]))
    VAT_amount = subtotal * (float(replacement_2.get("VAT-rate")[i])/100)
    replacement_2["VAT-amount"].append(VAT_amount)
    total += VAT_amount + subtotal
with open("invoices.csv", 'a') as invoice_file:
    invoice_file.write(str(new_id) + "," + str(date) + "," + str(time) + "," + str(company.lower()) + "," + str(total) + "\n")

print(replacement_2.get("DESCRIPTION")[1])
invoice_generator()







#VAT_amount = float(int(Number) * (VAT_rate * float(Amount)))
