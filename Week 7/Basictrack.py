import pandas as pd
import datetime
import openpyxl
from openpyxl.utils.cell import get_column_letter


def invoice_generator():
    """
    inputs:
    company             date
    address             new_id

    output:
    Excel file named:
    "company" + ".xlsx"
    """

    wb = openpyxl.load_workbook(r'Invoice_Template.xlsx')
    sheet = wb.active
    number_rows = sheet.max_row
    number_columns = sheet.max_column

    replacement1 = {
        "company": company,
        "date": date,
        "address": address,
        "id": "#" + str(new_id)
    }

    replacement2 = {
        "DESCRIPTION": list_description,
        "QTY": list_number,
        "UNIT PRICE": list_amount,
        "VAT-rate": list_VAT_rate,
        "VAT-amount": list_VAT_amount,
    }

    # find all replacement keys in excel file and change them with key value:
    for i in range(number_columns):
        for k in range(number_rows):
            cell = str(sheet[get_column_letter(i+1)+str(k+1)].value)
            for key in replacement1.keys():
                if str(cell) == key:
                    new_cell2 = replacement1.get(key)
                    sheet[get_column_letter(i+1)+str(k+1)] = str(new_cell2)

    # Append inputs to excel file
    # Excel file can be changed in accordance to current_row and current_column
    current_row = 16
    current_column = 2
    for eachRow in replacement2.keys():
        for i in range(len(list_description)):
            new_cell = replacement2.get(eachRow)[i]
            sheet.cell(row=current_row, column=current_column).value = new_cell
            current_row += 1
        current_column += 1
        current_row -= int(len(list_description))
    wb.save(str(company) + "#" + str(new_id) + ".xlsx")
    return replacement2


"""  
1. Create pandas dataframe, used to acces information from
company_address.csv and invoices.csv
2. Generate ID (need to be changed)
3. Ask for input (company)
4. Find address of company
"""

# Load in data as pandas DataFrame:
company_address = pd.read_csv("company_address.csv", delimiter=",")
df_address = pd.DataFrame(company_address)
invoices = pd.read_csv("invoices.csv", delimiter=",")
df_invoices = pd.DataFrame(invoices)

# generate ID:
if len(df_invoices) > 1:
    new_id = df_invoices.iloc[-1, 0] + 1
    print(new_id)
else:
    new_id = 1

# Check if company exists
company = input("Company: ")
date = datetime.datetime.today().strftime('%d-%m-%Y')
time = datetime.datetime.today().strftime('%H:%M:%S')

if company not in df_address['Company'].values:
    address = input("Addres: ")
    new_company = {
        "Company": company,
        "Address": address
    }
    with open('company_address.csv', 'a') as company_file:
        company_file.write(new_company.get("Company") + "," + new_company.get("Address") + "\n")
else:
    find_address = df_address.loc[df_address['Company'] == company]
    address = find_address.iloc[0, 1]

"""
1. Ask for input
2. Store input in lists
3. Add information to invoices.csv
"""

all_items = input("Amount of items: ")
list_description = []
list_number = []
list_amount = []
list_VAT_rate = []
list_VAT_amount = []
total = 0

for item in range(int(all_items)):
    Company = company
    Description = input("Product description: ")
    Number = float(input("Quantity: "))
    Amount = float(input("Unit price: "))
    VAT_rate = (int(input("VAT-rate: "))/100)
    VAT_amount = float(int(Number) * (VAT_rate * float(Amount)))
    list_description.append(Description)
    list_number.append(Number)
    list_amount.append(Amount)
    list_VAT_rate.append(VAT_rate)
    list_VAT_amount.append(VAT_amount)
    total += float((Number * Amount) + VAT_amount)

with open("invoices.csv", 'a') as invoice_file:
    invoice_file.write(str(new_id) + "," + str(date) + "," + str(time) + "," + str(company.lower()) + "," + str(total) + "\n")

invoice_generator()
